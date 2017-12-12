#!/usr/bin/python3
import pymysql
import openpyxl
from pprint import pprint
import sys
# Connect to the database
#------------------------------------------------
#This is Function section
#This getConnection() will connect to MysqlDb
#------------------------------------------------
def getConnection():
    connection = pymysql.connect(host='localhost',
                             user='root',
                             password='phani123',
                             db='lotuic',
                             charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)
    return connection
if __name__=='__main__':
    conn=getConnection()
    try:
        cursor=conn.cursor()
        wb = openpyxl.load_workbook('worksheet.xlsx')
        print (type(wb))
        print (wb.get_sheet_names())
        sheet=wb.get_sheet_by_name('Sheet1')
        max_row=sheet.max_row+1
        print (sheet.max_row)
        max_column=sheet.max_column+1
        print (max_column)
        header_dict={}
        excel_dict={}
        ##get first row headers
        for i in range(1,max_column):
            hdrn=sheet.cell(row=1,column=i).value
            print (hdrn)
            header_dict[i]=hdrn
            print (header_dict)
        ##from second row onwords
        for rn in range(2,max_row):
            excel_dict[rn]={}
            for cn in range(1,max_column):
                col_val=sheet.cell(row=rn,column=cn).value
                #print (col_val) 
                if rn==10:
                    sheet.cell(row=rn,column=cn).value="Phanindra"
                excel_dict[rn][header_dict[cn]]=col_val
        wb.save('worksheet.xlsx')
        pprint (excel_dict)
        for r in excel_dict:
            print (r)
            fn = excel_dict[r]['Category'];
            ln = excel_dict[r]['Description'];
            cy = excel_dict[r]['Function'];
            query = "insert into xl(Category,Function,Description) values('%s','%s','%s')" %(fn,ln,cy);
            print(query);
            #Execute mysql query.
            cursor.execute(query);
            conn.commit();

    except:
        print (sys.exc_info()[1])  
    finally:
        conn.close()
